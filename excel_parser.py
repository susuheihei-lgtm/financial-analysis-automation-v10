"""
Excelファイルからstock_data.json形式のデータを抽出するパーサー
シートが不足していても利用可能なデータだけで分析を行う
.xls（旧形式）および日本語ラベルの縦型レイアウトにも対応
ファジーマッチングによる柔軟なラベル検出機能搭載
"""
import os
import re
import unicodedata
from difflib import SequenceMatcher

import openpyxl
try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False


# ---------- xlrd → openpyxl 互換アダプタ ----------

class _XlrdCellAdapter:
    """xlrdのセル値をopenpyxlのcell.valueインターフェースで返す"""
    def __init__(self, value):
        self.value = value


class _XlrdSheetAdapter:
    """xlrdのシートをopenpyxlのワークシート互換で返す"""
    def __init__(self, xlrd_sheet):
        self._sheet = xlrd_sheet
        self.max_row = xlrd_sheet.nrows
        self.max_column = xlrd_sheet.ncols
        self.title = xlrd_sheet.name

    def cell(self, row, column):
        try:
            v = self._sheet.cell_value(row - 1, column - 1)
            if v == '':
                v = None
            return _XlrdCellAdapter(v)
        except IndexError:
            return _XlrdCellAdapter(None)


class _XlrdWorkbookAdapter:
    """xlrdのワークブックをopenpyxl互換で返す"""
    def __init__(self, xlrd_wb):
        self._wb = xlrd_wb
        self.sheetnames = xlrd_wb.sheet_names()

    def __getitem__(self, name):
        return _XlrdSheetAdapter(self._wb.sheet_by_name(name))


def _load_workbook(filepath):
    """拡張子に応じてopenpyxlまたはxlrdでワークブックを開く"""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.xls':
        if not HAS_XLRD:
            raise ImportError(
                '.xlsファイルの読み込みにはxlrdが必要です。'
                'pip install xlrd でインストールしてください。'
            )
        wb = xlrd.open_workbook(filepath)
        return _XlrdWorkbookAdapter(wb)
    else:
        return openpyxl.load_workbook(filepath, data_only=True)


# ---------- 統一シノニム辞書 ----------

METRIC_SYNONYMS = {
    # --- Income Statement ---
    'revenue': {
        'exact': ['Revenue', 'Total Revenue', 'Net Revenue', 'Sales',
                  '売上高', '収益'],
        'keywords': ['revenue', 'sales', 'net sales', '売上', '収益'],
        'anti_keywords': ['cost of revenue', 'cost of sales'],
        'value_type': 'large_number',
    },
    'cogs': {
        'exact': ['Cost of Revenue', 'Cost of Goods Sold', 'COGS',
                  '売上原価'],
        'keywords': ['cost of revenue', 'cost of goods', 'cogs', '売上原価'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },
    'op_income': {
        'exact': ['Operating Income', 'Operating Profit', 'EBIT',
                  '営業利益'],
        'keywords': ['operating income', 'operating profit', 'ebit', '営業利益'],
        'anti_keywords': ['non-operating', 'non operating'],
        'value_type': 'large_number',
    },
    'net_income': {
        'exact': ['Net Income', 'Net Profit', 'Net Earnings',
                  '純利益', '当期純利益'],
        'keywords': ['net income', 'net profit', 'net earnings', '純利益', '当期純利益'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },
    'eps': {
        'exact': ['EPS (Basic)', 'EPS', 'Earnings Per Share',
                  'EPS', '一株益', '1株益'],
        'keywords': ['eps', 'earnings per share', '一株益', '1株益'],
        'anti_keywords': ['diluted'],
        'value_type': 'per_share',
    },
    'op_margin': {
        'exact': ['Operating Margin', '営業利益率'],
        'keywords': ['operating margin', '営業利益率'],
        'anti_keywords': [],
        'value_type': 'percentage',
    },
    'ebitda_margin': {
        'exact': ['EBITDA Margin', 'EBITDAマージン'],
        'keywords': ['ebitda margin', 'ebitdaマージン'],
        'anti_keywords': [],
        'value_type': 'percentage',
    },
    'ebitda': {
        'exact': ['EBITDA'],
        'keywords': ['ebitda'],
        'anti_keywords': ['margin', 'ebitda margin'],
        'value_type': 'large_number',
    },
    'sga': {
        'exact': ['Selling, General & Admin', 'SG&A', 'SGA', '販管費'],
        'keywords': ['selling general', 'sg&a', 'sga', '販管費'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },
    'da': {
        'exact': ['Depreciation & Amortization', 'D&A', 'Depreciation',
                  '減価償却費'],
        'keywords': ['depreciation', 'amortization', 'd&a', '減価償却'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },
    'interest_exp': {
        'exact': ['Interest Expense / Income', 'Interest Expense',
                  'Net Interest Income', '支払利息'],
        'keywords': ['interest expense', 'interest income', 'net interest', '支払利息'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },
    'other_exp': {
        'exact': ['Other Expense / Income', 'Other Income/Expense',
                  'Non-Operating Income', '営業外損益'],
        'keywords': ['other expense', 'other income', 'non-operating', '営業外'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },
    'pretax_income': {
        'exact': ['Pretax Income', 'Pre-Tax Income', 'Income Before Tax',
                  '経常利益', '税引前利益'],
        'keywords': ['pretax', 'pre-tax', 'before tax', '経常利益', '税引前'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },
    'ordinary_income': {
        'exact': ['経常利益'],
        'keywords': ['経常利益'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },
    'income_tax': {
        'exact': ['Income Tax', 'Tax Provision', 'Provision for Income Taxes',
                  '法人税'],
        'keywords': ['income tax', 'tax provision', '法人税'],
        'anti_keywords': ['effective tax rate', 'before tax'],
        'value_type': 'large_number',
    },
    'eff_tax_rate': {
        'exact': ['Effective Tax Rate', '実効税率'],
        'keywords': ['effective tax rate', '実効税率'],
        'anti_keywords': [],
        'value_type': 'percentage',
    },
    'dates': {
        'exact': ['Date', 'Year Ending', 'Fiscal Year', 'Period',
                  '年度', '決算期', '決算年度'],
        'keywords': ['date', 'year', 'fiscal', 'period', '年度', '決算'],
        'anti_keywords': [],
        'value_type': None,
    },

    # --- Cash Flow ---
    'fcf': {
        'exact': ['Free Cash Flow', 'FCF', 'フリーCF'],
        'keywords': ['free cash flow', 'fcf', 'フリーcf'],
        'anti_keywords': ['debt/fcf'],
        'value_type': 'large_number',
    },
    'ocf': {
        'exact': ['Operating Cash Flow', 'Cash from Operations', '営業CF'],
        'keywords': ['operating cash flow', 'cash from operations', '営業cf'],
        'anti_keywords': ['margin'],
        'value_type': 'large_number',
    },
    'capex': {
        'exact': ['Capital Expenditures', 'Capex', 'CapEx', '設備投資'],
        'keywords': ['capital expenditure', 'capex', '設備投資'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },
    'investing_cf': {
        'exact': ['Investing Cash Flow', 'Cash from Investing', '投資CF'],
        'keywords': ['investing cash flow', 'cash from investing', '投資cf'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },
    'financing_cf': {
        'exact': ['Financing Cash Flow', 'Cash from Financing', '財務CF'],
        'keywords': ['financing cash flow', 'cash from financing', '財務cf'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },

    # --- Balance Sheet ---
    'total_assets': {
        'exact': ['Total Assets', '総資産'],
        'keywords': ['total assets', '総資産'],
        'anti_keywords': ['current assets', 'long-term assets'],
        'value_type': 'large_number',
    },
    'total_equity': {
        'exact': ['Shareholders Equity', "Shareholders' Equity",
                  'Total Equity', 'Stockholders Equity',
                  '株主資本', '自己資本', '純資産'],
        'keywords': ['shareholders equity', 'stockholders equity', 'total equity',
                     '株主資本', '自己資本', '純資産'],
        'anti_keywords': ['debt/equity', 'return on equity'],
        'value_type': 'large_number',
    },
    'total_debt': {
        'exact': ['Total Debt', 'Long-Term Debt', '有利子負債'],
        'keywords': ['total debt', 'long-term debt', '有利子負債'],
        'anti_keywords': ['net debt', 'debt/'],
        'value_type': 'large_number',
    },
    'receivables': {
        'exact': ['Receivables', 'Accounts Receivable', 'Trade Receivables',
                  '売上債権'],
        'keywords': ['receivables', 'accounts receivable', '売上債権'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },
    'inventory': {
        'exact': ['Inventory', 'Inventories', '棚卸資産'],
        'keywords': ['inventory', 'inventories', '棚卸資産'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },
    'payables': {
        'exact': ['Accounts Payable', 'Trade Payables', '買掛金'],
        'keywords': ['accounts payable', 'trade payable', '買掛金'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },
    'current_assets': {
        'exact': ['Total Current Assets', 'Current Assets', '流動資産'],
        'keywords': ['current assets', '流動資産'],
        'anti_keywords': ['non-current', 'long-term'],
        'value_type': 'large_number',
    },
    'current_liab': {
        'exact': ['Total Current Liabilities', 'Current Liabilities', '流動負債'],
        'keywords': ['current liabilities', '流動負債'],
        'anti_keywords': ['non-current'],
        'value_type': 'large_number',
    },
    'cash': {
        'exact': ['Cash & Cash Equivalents', 'Cash and Equivalents', 'Cash',
                  '現金同等物', '現金及び現金同等物'],
        'keywords': ['cash', 'cash equivalents', '現金'],
        'anti_keywords': ['net cash', 'free cash'],
        'value_type': 'large_number',
    },
    'fixed_assets': {
        'exact': ['Property, Plant & Equipment', 'PP&E', 'Fixed Assets',
                  '有形固定資産'],
        'keywords': ['property plant', 'pp&e', 'fixed assets', '有形固定資産'],
        'anti_keywords': ['intangible'],
        'value_type': 'large_number',
    },
    'intangibles': {
        'exact': ['Goodwill and Intangibles', 'Intangible Assets', 'Goodwill',
                  'のれん', '無形固定資産'],
        'keywords': ['goodwill', 'intangible', 'のれん', '無形固定資産'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },
    'net_debt': {
        'exact': ['Net Cash (Debt)', 'Net Debt', 'ネットキャッシュ'],
        'keywords': ['net cash', 'net debt', 'ネットキャッシュ'],
        'anti_keywords': ['net debt/ebitda'],
        'value_type': 'large_number',
    },
    'long_term_assets': {
        'exact': ['Total Long-Term Assets', 'Non-Current Assets', '固定資産'],
        'keywords': ['long-term assets', 'non-current assets', '固定資産'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },
    'retained_earnings': {
        'exact': ['Retained Earnings', '利益剰余金'],
        'keywords': ['retained earnings', '利益剰余金'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },
    'short_term_debt': {
        'exact': ['Short-Term Debt', '短期借入金'],
        'keywords': ['short-term debt', 'short term debt', '短期借入金'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },
    'long_term_debt': {
        'exact': ['Long-Term Debt', '長期借入金'],
        'keywords': ['long-term debt', 'long term debt', '長期借入金'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },
    'net_assets': {
        'exact': ['Net Assets', '純資産'],
        'keywords': ['net assets', '純資産'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },

    # --- Ratios ---
    'pe_ratio': {
        'exact': ['PE Ratio', 'P/E Ratio', 'Price/Earnings', 'PER'],
        'keywords': ['pe ratio', 'p/e', 'price/earnings', 'per'],
        'anti_keywords': [],
        'value_type': 'ratio',
    },
    'pb_ratio': {
        'exact': ['PB Ratio', 'P/B Ratio', 'Price/Book', 'PBR'],
        'keywords': ['pb ratio', 'p/b', 'price/book', 'pbr'],
        'anti_keywords': [],
        'value_type': 'ratio',
    },
    'ev': {
        'exact': ['Enterprise Value', 'EV'],
        'keywords': ['enterprise value', 'ev'],
        'anti_keywords': ['ev/ebitda'],
        'value_type': 'large_number',
    },
    'roe': {
        'exact': ['Return on Equity (ROE)', 'ROE', 'Return on Equity'],
        'keywords': ['return on equity', 'roe'],
        'anti_keywords': [],
        'value_type': 'percentage',
    },
    'roa': {
        'exact': ['Return on Assets (ROA)', 'ROA', 'Return on Assets'],
        'keywords': ['return on assets', 'roa'],
        'anti_keywords': [],
        'value_type': 'percentage',
    },
    'roic': {
        'exact': ['Return on Invested Capital (ROIC)', 'ROIC'],
        'keywords': ['return on invested capital', 'roic'],
        'anti_keywords': [],
        'value_type': 'percentage',
    },
    'current_ratio': {
        'exact': ['Current Ratio', '流動比率'],
        'keywords': ['current ratio', '流動比率'],
        'anti_keywords': [],
        'value_type': 'ratio',
    },
    'quick_ratio': {
        'exact': ['Quick Ratio', '当座比率'],
        'keywords': ['quick ratio', '当座比率'],
        'anti_keywords': [],
        'value_type': 'ratio',
    },
    'debt_fcf': {
        'exact': ['Debt/FCF'],
        'keywords': ['debt/fcf', 'debt fcf'],
        'anti_keywords': [],
        'value_type': 'ratio',
    },
    'debt_ebitda': {
        'exact': ['Debt/EBITDA'],
        'keywords': ['debt/ebitda', 'debt ebitda'],
        'anti_keywords': ['net debt/ebitda'],
        'value_type': 'ratio',
    },
    'nd_ebitda': {
        'exact': ['Net Debt/EBITDA'],
        'keywords': ['net debt/ebitda', 'net debt ebitda'],
        'anti_keywords': [],
        'value_type': 'ratio',
    },
    'debt_equity': {
        'exact': ['Debt/Equity', 'D/E Ratio'],
        'keywords': ['debt/equity', 'debt equity', 'd/e ratio'],
        'anti_keywords': [],
        'value_type': 'ratio',
    },
    'dividend_yield': {
        'exact': ['Dividend Yield', '配当利回り'],
        'keywords': ['dividend yield', '配当利回り'],
        'anti_keywords': [],
        'value_type': 'percentage',
    },
    'payout_ratio': {
        'exact': ['Payout Ratio', '配当性向'],
        'keywords': ['payout ratio', '配当性向'],
        'anti_keywords': ['total return'],
        'value_type': 'percentage',
    },

    # --- JP-specific ---
    'ocf_margin_pct': {
        'exact': ['営業CFマージン'],
        'keywords': ['営業cfマージン', '営業cf margin'],
        'anti_keywords': [],
        'value_type': 'percentage',
    },
    'bps': {
        'exact': ['BPS', '一株純資産', '1株純資産'],
        'keywords': ['bps', 'book value per share', '一株純資産', '1株純資産'],
        'anti_keywords': [],
        'value_type': 'per_share',
    },
    'equity_ratio_pct': {
        'exact': ['Equity Ratio', '自己資本比率'],
        'keywords': ['equity ratio', '自己資本比率'],
        'anti_keywords': [],
        'value_type': 'percentage',
    },
    'dividend_per_share': {
        'exact': ['Dividend Per Share', '一株配当', '1株配当'],
        'keywords': ['dividend per share', '一株配当', '1株配当'],
        'anti_keywords': [],
        'value_type': 'per_share',
    },
    'payout_ratio_pct': {
        'exact': ['Payout Ratio', '配当性向'],
        'keywords': ['payout ratio', '配当性向'],
        'anti_keywords': [],
        'value_type': 'percentage',
    },
    'total_return_ratio': {
        'exact': ['Total Shareholder Return', '総還元性向'],
        'keywords': ['total return ratio', '総還元性向'],
        'anti_keywords': [],
        'value_type': 'percentage',
    },
    'dividend_total': {
        'exact': ['剰余金の配当'],
        'keywords': ['剰余金の配当'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },
    'buyback': {
        'exact': ['Share Buyback', '自社株買い'],
        'keywords': ['share buyback', 'stock buyback', '自社株買い'],
        'anti_keywords': [],
        'value_type': 'large_number',
    },
    'doe': {
        'exact': ['DOE', '純資産配当率'],
        'keywords': ['doe', '純資産配当率'],
        'anti_keywords': [],
        'value_type': 'percentage',
    },
}

# ファジーマッチ閾値
_MATCH_THRESHOLD = 0.65


# ---------- ラベル正規化 ----------

def _normalize_label(text):
    """ラベルを正規化: Unicode正規化、小文字化、空白正規化、記号除去"""
    if text is None:
        return ''
    s = str(text).strip()
    # Unicode正規化（全角→半角）
    s = unicodedata.normalize('NFKC', s)
    s = s.lower()
    # 括弧とその中身は残すが、記号類を正規化
    s = re.sub(r'[（）\(\)\[\]【】{}]', ' ', s)
    s = re.sub(r'[—―\-–/／\\＆]', ' ', s)
    # 連続空白を1つに
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def _tokenize(text):
    """正規化済みテキストをトークン集合に分割"""
    return set(text.split())


# ---------- スコアリング関数 ----------

def _fuzzy_match_label(cell_text, metric_key):
    """セルテキストと指標キーのマッチスコアを返す (0.0〜1.0)"""
    syn = METRIC_SYNONYMS.get(metric_key)
    if syn is None:
        return 0.0

    raw = str(cell_text).strip() if cell_text is not None else ''
    if not raw:
        return 0.0

    normalized = _normalize_label(raw)
    if not normalized:
        return 0.0

    # 1. anti_keyword拒否
    for anti in syn.get('anti_keywords', []):
        if anti and _normalize_label(anti) in normalized:
            return 0.0

    # 2. 完全一致（raw or normalized vs exact candidates）
    for exact in syn.get('exact', []):
        if raw == exact:
            return 1.0
        if normalized == _normalize_label(exact):
            return 1.0

    # 3. 部分文字列一致（keyword in normalized or normalized in keyword）
    best_partial = 0.0
    for kw in syn.get('keywords', []):
        nkw = _normalize_label(kw)
        if not nkw:
            continue
        if nkw in normalized:
            # キーワードがラベルに含まれている：カバレッジで重み付け
            coverage = len(nkw) / len(normalized)
            score = 0.7 + 0.2 * coverage  # 0.7〜0.9
            best_partial = max(best_partial, score)
        elif normalized in nkw:
            coverage = len(normalized) / len(nkw)
            score = 0.7 + 0.2 * coverage
            best_partial = max(best_partial, score)

    if best_partial >= 0.7:
        return best_partial

    # 4. トークン重複（Jaccard類似度）
    norm_tokens = _tokenize(normalized)
    best_jaccard = 0.0
    for kw in syn.get('keywords', []):
        kw_tokens = _tokenize(_normalize_label(kw))
        if not kw_tokens:
            continue
        intersection = norm_tokens & kw_tokens
        union = norm_tokens | kw_tokens
        if union:
            jaccard = len(intersection) / len(union)
            # 0.5〜0.7にスケーリング
            score = 0.5 + 0.2 * jaccard
            best_jaccard = max(best_jaccard, score)

    if best_jaccard >= _MATCH_THRESHOLD:
        return best_jaccard

    # 5. SequenceMatcher（最終手段）
    best_seq = 0.0
    for kw in syn.get('keywords', []):
        nkw = _normalize_label(kw)
        if not nkw:
            continue
        ratio = SequenceMatcher(None, normalized, nkw).ratio()
        # 0.4〜0.7にスケーリング
        score = 0.4 + 0.3 * ratio
        best_seq = max(best_seq, score)

    return best_seq if best_seq >= _MATCH_THRESHOLD else 0.0


# ---------- データ妥当性検証 ----------

def _validate_match(values, value_type):
    """ファジーマッチ後にデータの値域を検証"""
    if value_type is None:
        return True
    nums = [v for v in values if isinstance(v, (int, float))]
    if not nums:
        return True  # データなしは通す（後段で処理）

    if value_type == 'percentage':
        return all(-200 <= v <= 500 for v in nums)
    elif value_type == 'large_number':
        return any(abs(v) >= 1 for v in nums)
    elif value_type == 'ratio':
        return all(-100 <= v <= 100 for v in nums)
    elif value_type == 'per_share':
        return all(-10000 <= v <= 100000 for v in nums)
    return True


# ---------- ファジー行データ取得 ----------

def _get_row_data(ws, row_label):
    """指定ラベルの行データを取得（新しい順）— 完全一致版（scan_available_metrics用に残す）"""
    if ws is None:
        return []
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == row_label:
            vals = []
            for c in range(2, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    vals.append(v)
            return list(reversed(vals))  # 新しい順
    return []


def _fuzzy_get_row_data(ws, metric_key):
    """ファジーマッチで指標の行データを取得。
    Returns: (data_list, matched_label, score) — data_listは新しい順"""
    if ws is None:
        return [], None, 0.0

    best_row = None
    best_score = 0.0
    best_label = None

    for r in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=1).value
        if cell_val is None:
            continue
        score = _fuzzy_match_label(cell_val, metric_key)
        if score > best_score:
            best_score = score
            best_row = r
            best_label = str(cell_val).strip()

    if best_score < _MATCH_THRESHOLD:
        return [], None, 0.0

    # 行データ取得
    vals = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=best_row, column=c).value
        if v is not None:
            vals.append(v)
    data = list(reversed(vals))  # 新しい順

    # 妥当性検証
    syn = METRIC_SYNONYMS.get(metric_key, {})
    vtype = syn.get('value_type')
    if best_score < 1.0 and not _validate_match(data, vtype):
        return [], None, 0.0

    return data, best_label, best_score


def _fuzzy_find_all_metrics(ws, metric_keys):
    """複数指標を一括検索（高速化：ラベルを1回だけスキャン）。
    Returns: dict of metric_key -> (data_list, matched_label, score)"""
    results = {k: ([], None, 0.0) for k in metric_keys}
    if ws is None:
        return results

    # 全行のラベルを1回スキャン
    row_labels = {}  # row_num -> cell_value
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None:
            row_labels[r] = v

    # 各指標に対してベストマッチを探す
    for metric_key in metric_keys:
        best_row = None
        best_score = 0.0
        best_label = None

        for r, cell_val in row_labels.items():
            score = _fuzzy_match_label(cell_val, metric_key)
            if score > best_score:
                best_score = score
                best_row = r
                best_label = str(cell_val).strip()

        if best_score < _MATCH_THRESHOLD:
            continue

        vals = []
        for c in range(2, ws.max_column + 1):
            v = ws.cell(row=best_row, column=c).value
            if v is not None:
                vals.append(v)
        data = list(reversed(vals))

        syn = METRIC_SYNONYMS.get(metric_key, {})
        vtype = syn.get('value_type')
        if best_score < 1.0 and not _validate_match(data, vtype):
            continue

        results[metric_key] = (data, best_label, best_score)

    return results


# ---------- 日本語ラベル ファジーマッチ ----------

def _fuzzy_match_jp_label(label):
    """日本語ラベルをMETRIC_SYNONYMSの全キーに対してマッチし、最良一致のキーを返す"""
    if label is None:
        return None
    best_key = None
    best_score = 0.0
    for metric_key in METRIC_SYNONYMS:
        score = _fuzzy_match_label(label, metric_key)
        if score > best_score:
            best_score = score
            best_key = metric_key
    if best_score >= _MATCH_THRESHOLD:
        return best_key
    return None


# ---------- 既存のヘルパー関数 ----------

def _safe_get(lst, idx, default=None):
    return lst[idx] if idx < len(lst) else default


def _find_sheet(wb, candidates):
    """候補名リストからシートを探す。正規化してファジーに見つからなければNone"""
    # 完全一致
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    # 正規化比較
    norm_sheets = {_normalize_label(s): s for s in wb.sheetnames}
    for name in candidates:
        norm_name = _normalize_label(name)
        # 完全一致（正規化後）
        if norm_name in norm_sheets:
            return wb[norm_sheets[norm_name]]
        # 部分一致
        for norm_key, real_name in norm_sheets.items():
            if norm_name in norm_key or norm_key in norm_name:
                return wb[real_name]
    return None


# ---------- 日本語縦型レイアウト検出・パーサー ----------

# セクションヘッダー
_JP_SECTION_HEADERS = {'業績', '財務', 'CF', 'キャッシュフロー', '配当', '株価指標'}
_JP_SECTION_HEADERS_NORMALIZED = {_normalize_label(h) for h in _JP_SECTION_HEADERS}


def _is_jp_section_header(text):
    """セクションヘッダーかどうかをファジーに判定"""
    if text is None:
        return False
    raw = str(text).strip()
    if raw in _JP_SECTION_HEADERS:
        return True
    norm = _normalize_label(raw)
    return norm in _JP_SECTION_HEADERS_NORMALIZED


def _is_japanese_vertical_layout(wb):
    """日本語の縦型レイアウト（1シートに業績・財務・CF等が並ぶ）かどうかを判定"""
    if len(wb.sheetnames) > 4:
        return False
    ws = wb[wb.sheetnames[0]]
    found_sections = set()
    for r in range(1, min(ws.max_row + 1, 50)):
        val = ws.cell(row=r, column=1).value
        if _is_jp_section_header(val):
            found_sections.add(_normalize_label(str(val).strip()))
    return len(found_sections) >= 2


def _parse_numeric(v):
    """値を数値に変換。'-'や空文字、'（予想）'等はNoneにする"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(',', '')
    if s in ('-', '－', '', '―', 'N/A', 'n/a'):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _parse_japanese_vertical(wb, currency='JPY'):
    """日本語縦型レイアウトのExcelをパースしてanalyzer用のdictとts_dataを返す"""
    ws = wb[wb.sheetnames[0]]

    # 全行を読み込み
    rows = []
    for r in range(1, ws.max_row + 1):
        row = []
        for c in range(1, ws.max_column + 1):
            row.append(ws.cell(row=r, column=c).value)
        rows.append(row)

    # 会社名の検出（1行目）
    company_name = str(rows[0][0]).strip() if rows and rows[0][0] else ''

    # セクションごとにデータを読み取る
    raw_data = {}  # key -> [値のリスト（古い→新しい順）]
    dates_by_section = {}
    current_section = None
    header_row = None

    for i, row in enumerate(rows):
        first_cell = str(row[0]).strip() if row[0] is not None else ''

        # セクションヘッダーを検出（ファジー）
        if _is_jp_section_header(first_cell):
            current_section = first_cell
            header_row = None
            continue

        # セクション内のラベル行を検出
        if current_section and header_row is None:
            norm_first = _normalize_label(first_cell)
            if norm_first in ('年度', '決算期', '決算年度'):
                header_row = row
                continue
            continue

        # データ行を処理
        if current_section and header_row is not None:
            # 空行でセクション終了
            if not first_cell:
                current_section = None
                header_row = None
                continue

            # 予想行はスキップ
            last_cell = row[-1] if row else None
            if last_cell and '予想' in str(last_cell):
                continue

            # 年度を取得
            date_val = first_cell
            if current_section not in dates_by_section:
                dates_by_section[current_section] = []
            dates_by_section[current_section].append(date_val)

            # 各列のデータを対応するキーに格納（ファジーマッチ）
            for col_idx in range(1, len(header_row)):
                label = str(header_row[col_idx]).strip() if header_row[col_idx] else ''
                if not label or label == '':
                    continue
                key = _fuzzy_match_jp_label(label)
                if key is None:
                    continue
                val = _parse_numeric(row[col_idx] if col_idx < len(row) else None)
                if key not in raw_data:
                    raw_data[key] = []
                raw_data[key].append(val)

    # 日付を取得（業績セクションを優先）
    dates_raw = dates_by_section.get('業績', dates_by_section.get(
        list(dates_by_section.keys())[0] if dates_by_section else '', []))

    # データは古い→新しい順で並んでいるので、新しい順に反転
    for key in raw_data:
        raw_data[key] = list(reversed(raw_data[key]))
    dates_raw = list(reversed(dates_raw))

    # ヘルパー
    def g(key, idx):
        lst = raw_data.get(key, [])
        return lst[idx] if idx < len(lst) else None

    def g_list(key):
        return raw_data.get(key, [])

    # ROE/ROAは%値として直接入っている場合がある（10.86 = 10.86%）
    roe_list = g_list('roe')
    roa_list = g_list('roa')
    roe_now = g('roe', 0)
    roe_3y = g('roe', 2)
    roe_5y = g('roe', 4)
    roa_now = g('roa', 0)
    roa_3y = g('roa', 2)
    roa_5y = g('roa', 4)
    roe_growth = roe_now - roe_5y if (roe_now is not None and roe_5y is not None) else None

    # 自己資本比率（%値で直接入っている）
    equity_ratio = g('equity_ratio_pct', 0)
    equity_ratio_5y = g('equity_ratio_pct', 4)
    if equity_ratio is None and g('total_equity', 0) and g('total_assets', 0):
        equity_ratio = (g('total_equity', 0) / g('total_assets', 0)) * 100
    if equity_ratio_5y is None and g('total_equity', 4) and g('total_assets', 4):
        equity_ratio_5y = (g('total_equity', 4) / g('total_assets', 4)) * 100

    # 営業利益率の計算
    revenue = g_list('revenue')
    op_income = g_list('op_income')
    net_income = g_list('net_income')
    op_margin_vals = []
    if g_list('op_margin'):
        op_margin_vals = g_list('op_margin')[:5]
    elif g_list('op_margin_pct'):
        op_margin_vals = g_list('op_margin_pct')[:5]
    elif op_income and revenue:
        for i in range(min(5, len(op_income))):
            oi = g('op_income', i)
            rev = g('revenue', i)
            if oi is not None and rev and rev != 0:
                op_margin_vals.append(round(oi / rev * 100, 2))
            else:
                op_margin_vals.append(None)

    # FCF計算（営業CF - 設備投資の絶対値）
    ocf_list = g_list('ocf')
    capex_list = g_list('capex')
    fcf_list = g_list('fcf')
    if not fcf_list and ocf_list and capex_list:
        fcf_list = []
        for i in range(min(len(ocf_list), len(capex_list))):
            o = ocf_list[i]
            c = capex_list[i]
            if o is not None and c is not None:
                fcf_list.append(o + c)  # capexは通常マイナス値
            else:
                fcf_list.append(None)

    # 配当性向（%値）
    payout_pct = g('payout_ratio_pct', 0)
    if payout_pct is None:
        payout_pct = g('payout_ratio', 0)
    payout_pct_5y = g('payout_ratio_pct', 4)
    if payout_pct_5y is None:
        payout_pct_5y = g('payout_ratio', 4)

    # NOPAT
    nopat = g('op_income', 0) * 0.75 if g('op_income', 0) else None
    nopat_5y = g('op_income', 4) * 0.75 if g('op_income', 4) else None

    data = {
        "company": company_name,
        "ticker": "",
        "industry": "製造・サービス",

        "revenue": [g('revenue', i) for i in range(min(5, len(revenue)))],
        "fcf": [fcf_list[i] if i < len(fcf_list) else None for i in range(min(5, len(fcf_list)))],
        "eps": [g('eps', i) for i in range(min(5, len(g_list('eps'))))],

        "roe": [roe_now, roe_3y, roe_5y],
        "roe_growth_rate": roe_growth,
        "roa": [roa_now, roa_3y, roa_5y],

        "equity_ratio": equity_ratio,
        "equity_ratio_5y": equity_ratio_5y,
        "quick_ratio": None,
        "quick_ratio_5y": None,
        "current_ratio": None,
        "current_ratio_5y": None,

        "operating_cf": [g('ocf', i) for i in range(min(5, len(ocf_list)))],
        "investing_cf": [g('investing_cf', i) for i in range(min(5, len(g_list('investing_cf'))))],
        "financing_cf": [g('financing_cf', i) for i in range(min(5, len(g_list('financing_cf'))))],
        "op_margin": op_margin_vals,
        "ebitda_margin": None,
        "ebitda_margin_5y": None,

        "debt_fcf": None,
        "debt_fcf_5y": None,
        "nd_ebitda": None,
        "ev": None,
        "per": None,
        "per_5y": None,
        "pbr": None,
        "pbr_5y": None,

        "nopat": nopat,
        "nopat_5y": nopat_5y,
        "invested_capital": None,
        "invested_capital_5y": None,
        "wacc": None,

        "accounts_receivable": None,
        "accounts_receivable_5y": None,
        "inventory": None,
        "inventory_5y": None,
        "accounts_payable": None,
        "accounts_payable_5y": None,
        "cogs": None,
        "cogs_5y": None,
        "sga_ratio": None,
        "sga_ratio_5y": None,

        "total_assets": g('total_assets', 0),
        "total_assets_5y": g('total_assets', 4),
        "fixed_assets": None,
        "fixed_assets_5y": None,
        "tangible_fixed_assets": None,
        "tangible_fixed_assets_5y": None,
        "intangible_fixed_assets": None,
        "intangible_fixed_assets_5y": None,

        "net_income_val": g('net_income', 0),
        "net_income_val_5y": g('net_income', 4),
        "op_income_val": g('op_income', 0),
        "op_income_val_5y": g('op_income', 4),
        "interest_exp": None,
        "interest_exp_5y": None,
        "other_exp": None,
        "other_exp_5y": None,
        "pretax_income": None,
        "pretax_income_5y": None,
        "income_tax": None,
        "income_tax_5y": None,
        "total_equity": g('total_equity', 0),
        "total_equity_5y": g('total_equity', 4),

        "dividend_yield": None,
        "dividend_yield_5y": None,
        "payout_ratio": payout_pct,
        "payout_ratio_5y": payout_pct_5y,

        "d1_mgmt_change": "○",
        "d2_ownership": "○",
        "d3_esg": "○",
    }

    # 時系列データ（チャート用）
    date_strs = [str(d)[:4] if d else "" for d in dates_raw]
    investing_cf_list = g_list('investing_cf')
    financing_cf_list = g_list('financing_cf')
    eps_list = g_list('eps')

    ts_data = {
        "dates": date_strs,
        "revenue": list(revenue),
        "net_income": list(net_income),
        "fcf": list(fcf_list),
        "eps": list(eps_list),
        "ocf": list(ocf_list),
        "investing_cf": list(investing_cf_list),
        "financing_cf": list(financing_cf_list),
        "ebitda": [],
        "total_assets": list(g_list('total_assets')),
        "total_equity": list(g_list('total_equity')),
        "total_debt": list(g_list('total_debt')),
        "roe": list(roe_list),
        "roa": list(roa_list),
        "op_margin": list(op_margin_vals) + [None] * max(0, len(revenue) - len(op_margin_vals)),
        "quick_ratio": [],
        "current_ratio": [],
        "equity_ratio": list(g_list('equity_ratio_pct')),
        "ebitda_margin": [],
        "debt_fcf": [],
        "roic": [],
        "capex": list(capex_list),
        "sga": [],
        "da": [],
        "pe_ratio": [],
        "pb_ratio": [],
        "debt_ebitda": [],
        "nd_ebitda": [],
        "dividend_yield": [],
        "payout_ratio": list(g_list('payout_ratio_pct')),
    }

    # 営業利益率の時系列計算
    if not ts_data["op_margin"] or all(v is None for v in ts_data["op_margin"]):
        ts_data["op_margin"] = []
        for i in range(len(op_income)):
            oi = op_income[i] if i < len(op_income) else None
            rev = revenue[i] if i < len(revenue) else None
            if oi is not None and rev and rev != 0:
                ts_data["op_margin"].append(round(oi / rev * 100, 2))
            else:
                ts_data["op_margin"].append(None)

    # DuPont分解
    max_len = len(revenue) if revenue else 0
    total_assets_list = g_list('total_assets')
    total_equity_list = g_list('total_equity')
    net_margin_ts = []
    asset_turnover_ts = []
    fin_leverage_ts = []
    for i in range(max_len):
        ni = net_income[i] if i < len(net_income) else None
        rev = revenue[i] if i < len(revenue) else None
        ta = total_assets_list[i] if i < len(total_assets_list) else None
        eq = total_equity_list[i] if i < len(total_equity_list) else None
        if ni is not None and rev and rev != 0:
            net_margin_ts.append(round(ni / rev * 100, 2))
        else:
            net_margin_ts.append(None)
        if rev is not None and ta and ta != 0:
            asset_turnover_ts.append(round(rev / ta, 3))
        else:
            asset_turnover_ts.append(None)
        if ta is not None and eq and eq != 0:
            fin_leverage_ts.append(round(ta / eq, 3))
        else:
            fin_leverage_ts.append(None)

    ts_data["net_margin"] = net_margin_ts
    ts_data["asset_turnover"] = asset_turnover_ts
    ts_data["financial_leverage"] = fin_leverage_ts
    ts_data["interest_burden"] = []
    ts_data["nonop_burden"] = []
    ts_data["tax_burden"] = []

    # Metadata for frontend scaling
    ts_data["_source"] = "excel"
    ts_data["_unit_scale"] = "millions"  # Data is already in millions
    ts_data["_is_jpy"] = False  # Prevent double-scaling in smartFormat
    ts_data["_currency"] = currency

    return data, ts_data


# ---------- メインパース関数 ----------

def parse_excel(filepath, currency='JPY'):
    """Excelファイルをパースしてanalyzer用のdictを返す。
    .xls/.xlsx両対応。日本語縦型レイアウトも自動検出する。"""
    wb = _load_workbook(filepath)

    # 日本語縦型レイアウトの検出
    if _is_japanese_vertical_layout(wb):
        return _parse_japanese_vertical(wb, currency=currency)

    # --- 以下、従来の英語マルチシート形式 ---

    # シートの自動検出（複数の名前パターンに対応）
    inc = _find_sheet(wb, ['Income-Annual', 'Income Statement', 'Income', 'Export', 'income'])
    bs = _find_sheet(wb, ['Balance-Sheet-Annual', 'Balance Sheet', 'Balance', 'balance'])
    cf = _find_sheet(wb, ['Cash-Flow-Annual', 'Cash Flow', 'Cash Flow Statement', 'cashflow'])
    rat = _find_sheet(wb, ['Ratios-Annual', 'Ratios', 'Financial Ratios', 'ratios'])

    # もし全シートNoneなら、最初のシートをincとして使う
    if inc is None and bs is None and cf is None and rat is None:
        if wb.sheetnames:
            inc = wb[wb.sheetnames[0]]

    # 一括ファジー検索 — 各シートのラベルを1回だけスキャン
    inc_metrics = _fuzzy_find_all_metrics(inc, [
        'dates', 'revenue', 'cogs', 'op_income', 'net_income', 'eps',
        'op_margin', 'ebitda_margin', 'ebitda', 'sga', 'da',
        'interest_exp', 'other_exp', 'pretax_income', 'income_tax', 'eff_tax_rate',
    ])
    cf_metrics = _fuzzy_find_all_metrics(cf, [
        'fcf', 'ocf', 'capex', 'investing_cf', 'financing_cf',
    ])
    bs_metrics = _fuzzy_find_all_metrics(bs, [
        'total_assets', 'total_equity', 'total_debt', 'receivables',
        'inventory', 'payables', 'current_assets', 'current_liab',
        'cash', 'fixed_assets', 'intangibles', 'net_debt', 'long_term_assets',
    ])
    rat_metrics = _fuzzy_find_all_metrics(rat, [
        'pe_ratio', 'pb_ratio', 'ev', 'roe', 'roa', 'roic',
        'current_ratio', 'quick_ratio', 'debt_fcf', 'debt_ebitda',
        'nd_ebitda', 'debt_equity', 'dividend_yield', 'payout_ratio',
    ])

    # ヘルパー: メトリクスデータ取得
    def fm(metrics_dict, key):
        return metrics_dict[key][0] if key in metrics_dict else []

    # 日付（新しい順）
    dates_raw = fm(inc_metrics, 'dates')

    # 収益データ
    revenue = fm(inc_metrics, 'revenue')
    cogs_list = fm(inc_metrics, 'cogs')
    op_income = fm(inc_metrics, 'op_income')
    net_income = fm(inc_metrics, 'net_income')
    eps_list = fm(inc_metrics, 'eps')
    op_margin_list = fm(inc_metrics, 'op_margin')
    ebitda_margin_list = fm(inc_metrics, 'ebitda_margin')
    ebitda_list = fm(inc_metrics, 'ebitda')
    sga_list = fm(inc_metrics, 'sga')
    da_list = fm(inc_metrics, 'da')
    interest_exp_list = fm(inc_metrics, 'interest_exp')
    other_exp_list = fm(inc_metrics, 'other_exp')
    pretax_income_list = fm(inc_metrics, 'pretax_income')
    income_tax_list = fm(inc_metrics, 'income_tax')
    eff_tax_rate_list = fm(inc_metrics, 'eff_tax_rate')

    # キャッシュフロー
    fcf_list = fm(cf_metrics, 'fcf')
    ocf_list = fm(cf_metrics, 'ocf')
    capex_list = fm(cf_metrics, 'capex')
    investing_cf_list = fm(cf_metrics, 'investing_cf')
    financing_cf_list = fm(cf_metrics, 'financing_cf')

    # バランスシート
    total_assets_list = fm(bs_metrics, 'total_assets')
    total_equity_list = fm(bs_metrics, 'total_equity')
    total_debt_list = fm(bs_metrics, 'total_debt')
    receivables_list = fm(bs_metrics, 'receivables')
    inventory_list = fm(bs_metrics, 'inventory')
    payables_list = fm(bs_metrics, 'payables')
    current_assets_list = fm(bs_metrics, 'current_assets')
    current_liab_list = fm(bs_metrics, 'current_liab')
    cash_list = fm(bs_metrics, 'cash')
    fixed_assets_list = fm(bs_metrics, 'fixed_assets')
    intangibles_list = fm(bs_metrics, 'intangibles')
    net_debt_list = fm(bs_metrics, 'net_debt')
    long_term_assets_list = fm(bs_metrics, 'long_term_assets')

    # 比率
    pe_list = fm(rat_metrics, 'pe_ratio')
    pb_list = fm(rat_metrics, 'pb_ratio')
    ev_list = fm(rat_metrics, 'ev')
    roe_list = fm(rat_metrics, 'roe')
    roa_list = fm(rat_metrics, 'roa')
    current_ratio_list = fm(rat_metrics, 'current_ratio')
    quick_ratio_list = fm(rat_metrics, 'quick_ratio')
    debt_fcf_list = fm(rat_metrics, 'debt_fcf')
    debt_ebitda_list = fm(rat_metrics, 'debt_ebitda')
    nd_ebitda_list = fm(rat_metrics, 'nd_ebitda')
    roic_list = fm(rat_metrics, 'roic')
    debt_equity_list = fm(rat_metrics, 'debt_equity')
    dividend_yield_list = fm(rat_metrics, 'dividend_yield')
    payout_ratio_list = fm(rat_metrics, 'payout_ratio')

    # ROE/ROAがRatiosシートにないがIncome+BSから計算可能な場合
    if not roe_list and net_income and total_equity_list:
        roe_list = []
        max_len_roe = min(len(net_income), len(total_equity_list))
        for i in range(max_len_roe):
            ni = _safe_get(net_income, i)
            eq = _safe_get(total_equity_list, i)
            if ni is not None and eq and eq != 0:
                roe_list.append(ni / eq)
            else:
                roe_list.append(None)

    if not roa_list and net_income and total_assets_list:
        roa_list = []
        max_len_roa = min(len(net_income), len(total_assets_list))
        for i in range(max_len_roa):
            ni = _safe_get(net_income, i)
            ta = _safe_get(total_assets_list, i)
            if ni is not None and ta and ta != 0:
                roa_list.append(ni / ta)
            else:
                roa_list.append(None)

    # 値の取得ヘルパー
    def g(lst, idx):
        return _safe_get(lst, idx)

    # パーセント変換
    def to_pct(v):
        return v * 100 if v is not None else None

    # 自己資本比率の計算
    equity_ratio = None
    equity_ratio_5y = None
    de0 = g(debt_equity_list, 0)
    de4 = g(debt_equity_list, 4)
    if de0 is not None:
        equity_ratio = (1 / (1 + de0)) * 100
    elif g(total_equity_list, 0) and g(total_assets_list, 0):
        equity_ratio = (g(total_equity_list, 0) / g(total_assets_list, 0)) * 100
    if de4 is not None:
        equity_ratio_5y = (1 / (1 + de4)) * 100
    elif g(total_equity_list, 4) and g(total_assets_list, 4):
        equity_ratio_5y = (g(total_equity_list, 4) / g(total_assets_list, 4)) * 100

    # 当座比率
    quick_r = to_pct(g(quick_ratio_list, 0))
    quick_r_5y = to_pct(g(quick_ratio_list, 4))

    # 流動比率
    current_r = to_pct(g(current_ratio_list, 0))
    current_r_5y = to_pct(g(current_ratio_list, 4))

    # 営業利益率（%単位に変換）— 5年分
    op_margin_vals = [to_pct(g(op_margin_list, i)) for i in range(min(5, len(op_margin_list)))]
    if not op_margin_vals and op_income and revenue:
        op_margin_vals = []
        for i in range(min(5, len(op_income))):
            oi = g(op_income, i)
            rev = g(revenue, i)
            if oi is not None and rev and rev != 0:
                op_margin_vals.append(round(oi / rev * 100, 2))
            else:
                op_margin_vals.append(None)

    # EBITDAマージン
    ebitda_margin_5y = to_pct(g(ebitda_margin_list, 4))
    ebitda_margin_val = to_pct(g(ebitda_margin_list, 0))

    # ROE, ROA（%単位に変換）
    roe_now = to_pct(g(roe_list, 0))
    roe_3y = to_pct(g(roe_list, 2))
    roe_5y = to_pct(g(roe_list, 4))
    roa_now = to_pct(g(roa_list, 0))
    roa_3y = to_pct(g(roa_list, 2))
    roa_5y = to_pct(g(roa_list, 4))

    # ROE成長率
    roe_growth = roe_now - roe_5y if (roe_now is not None and roe_5y is not None) else None

    # NOPAT
    nopat = g(op_income, 0) * 0.75 if g(op_income, 0) else None
    nopat_5y = g(op_income, 4) * 0.75 if g(op_income, 4) else None

    # 投下資本
    def calc_ic(eq, debt, cash):
        if eq is not None and debt is not None and cash is not None:
            return eq + debt - cash
        return None

    ic = calc_ic(g(total_equity_list, 0), g(total_debt_list, 0), g(cash_list, 0))
    ic_5y = calc_ic(g(total_equity_list, 4), g(total_debt_list, 4), g(cash_list, 4))

    # WACC簡易推定
    wacc_val = None
    if g(total_equity_list, 0) and g(total_debt_list, 0) and g(total_assets_list, 0):
        d_ratio = g(total_debt_list, 0) / (g(total_equity_list, 0) + g(total_debt_list, 0))
        e_ratio = 1 - d_ratio
        cost_of_equity = 8.0
        cost_of_debt = 3.0
        tax_rate = 0.25
        wacc_val = e_ratio * cost_of_equity + d_ratio * cost_of_debt * (1 - tax_rate)

    # SGA比率
    sga_ratio = None
    sga_ratio_5y = None
    if g(sga_list, 0) and g(revenue, 0):
        sga_ratio = (g(sga_list, 0) / g(revenue, 0)) * 100
    if g(sga_list, 4) and g(revenue, 4):
        sga_ratio_5y = (g(sga_list, 4) / g(revenue, 4)) * 100

    data = {
        "company": "",
        "ticker": "",
        "industry": "製造・サービス",

        "revenue": [g(revenue, i) for i in range(min(5, len(revenue)))],
        "fcf": [g(fcf_list, i) for i in range(min(5, len(fcf_list)))],
        "eps": [g(eps_list, i) for i in range(min(5, len(eps_list)))],

        "roe": [roe_now, roe_3y, roe_5y],
        "roe_growth_rate": roe_growth,
        "roa": [roa_now, roa_3y, roa_5y],

        "equity_ratio": equity_ratio,
        "equity_ratio_5y": equity_ratio_5y,
        "quick_ratio": quick_r,
        "quick_ratio_5y": quick_r_5y,
        "current_ratio": current_r,
        "current_ratio_5y": current_r_5y,

        "operating_cf": [g(ocf_list, i) for i in range(min(5, len(ocf_list)))],
        "investing_cf": [g(investing_cf_list, i) for i in range(min(5, len(investing_cf_list)))],
        "financing_cf": [g(financing_cf_list, i) for i in range(min(5, len(financing_cf_list)))],
        "op_margin": op_margin_vals,
        "ebitda_margin": ebitda_margin_val,
        "ebitda_margin_5y": ebitda_margin_5y,

        "debt_fcf": g(debt_fcf_list, 0),
        "debt_fcf_5y": g(debt_fcf_list, 4),
        "nd_ebitda": g(nd_ebitda_list, 0),
        "ev": g(ev_list, 0),
        "per": g(pe_list, 0),
        "per_5y": g(pe_list, 4),
        "pbr": g(pb_list, 0),
        "pbr_5y": g(pb_list, 4),

        "nopat": nopat,
        "nopat_5y": nopat_5y,
        "invested_capital": ic,
        "invested_capital_5y": ic_5y,
        "wacc": wacc_val,

        "accounts_receivable": g(receivables_list, 0),
        "accounts_receivable_5y": g(receivables_list, 4),
        "inventory": g(inventory_list, 0),
        "inventory_5y": g(inventory_list, 4),
        "accounts_payable": g(payables_list, 0),
        "accounts_payable_5y": g(payables_list, 4),
        "cogs": g(cogs_list, 0),
        "cogs_5y": g(cogs_list, 4),
        "sga_ratio": sga_ratio,
        "sga_ratio_5y": sga_ratio_5y,

        "total_assets": g(total_assets_list, 0),
        "total_assets_5y": g(total_assets_list, 4),
        "fixed_assets": g(fixed_assets_list, 0),
        "fixed_assets_5y": g(fixed_assets_list, 4),
        "tangible_fixed_assets": g(fixed_assets_list, 0),
        "tangible_fixed_assets_5y": g(fixed_assets_list, 4),
        "intangible_fixed_assets": g(intangibles_list, 0),
        "intangible_fixed_assets_5y": g(intangibles_list, 4),

        "net_income_val": g(net_income, 0),
        "net_income_val_5y": g(net_income, 4),
        "op_income_val": g(op_income, 0),
        "op_income_val_5y": g(op_income, 4),
        "interest_exp": g(interest_exp_list, 0),
        "interest_exp_5y": g(interest_exp_list, 4),
        "other_exp": g(other_exp_list, 0),
        "other_exp_5y": g(other_exp_list, 4),
        "pretax_income": g(pretax_income_list, 0),
        "pretax_income_5y": g(pretax_income_list, 4),
        "income_tax": g(income_tax_list, 0),
        "income_tax_5y": g(income_tax_list, 4),
        "total_equity": g(total_equity_list, 0),
        "total_equity_5y": g(total_equity_list, 4),

        "dividend_yield": to_pct(g(dividend_yield_list, 0)),
        "dividend_yield_5y": to_pct(g(dividend_yield_list, 4)),
        "payout_ratio": to_pct(g(payout_ratio_list, 0)),
        "payout_ratio_5y": to_pct(g(payout_ratio_list, 4)),

        "d1_mgmt_change": "○",
        "d2_ownership": "○",
        "d3_esg": "○",
    }

    # 時系列データ（チャート用）
    ts_data = {
        "dates": [str(d)[:4] if d else "" for d in dates_raw],
        "revenue": [g(revenue, i) for i in range(len(revenue))],
        "net_income": [g(net_income, i) for i in range(len(net_income))],
        "fcf": [g(fcf_list, i) for i in range(len(fcf_list))],
        "eps": [g(eps_list, i) for i in range(len(eps_list))],
        "ocf": [g(ocf_list, i) for i in range(len(ocf_list))],
        "investing_cf": [g(investing_cf_list, i) for i in range(len(investing_cf_list))],
        "financing_cf": [g(financing_cf_list, i) for i in range(len(financing_cf_list))],
        "ebitda": [g(ebitda_list, i) for i in range(len(ebitda_list))],
        "total_assets": [g(total_assets_list, i) for i in range(len(total_assets_list))],
        "total_equity": [g(total_equity_list, i) for i in range(len(total_equity_list))],
        "total_debt": [g(total_debt_list, i) for i in range(len(total_debt_list))],
        "roe": [to_pct(g(roe_list, i)) for i in range(len(roe_list))],
        "roa": [to_pct(g(roa_list, i)) for i in range(len(roa_list))],
        "op_margin": [to_pct(g(op_margin_list, i)) for i in range(len(op_margin_list))],
        "quick_ratio": [to_pct(g(quick_ratio_list, i)) for i in range(len(quick_ratio_list))],
        "current_ratio": [to_pct(g(current_ratio_list, i)) for i in range(len(current_ratio_list))],
        "equity_ratio": [((1 / (1 + g(debt_equity_list, i))) * 100) if (g(debt_equity_list, i) is not None) else ((g(total_equity_list, i) / g(total_assets_list, i) * 100) if (g(total_equity_list, i) and g(total_assets_list, i)) else None) for i in range(max(len(debt_equity_list), len(total_equity_list)))],
        "ebitda_margin": [to_pct(g(ebitda_margin_list, i)) for i in range(len(ebitda_margin_list))],
        "debt_fcf": [g(debt_fcf_list, i) for i in range(len(debt_fcf_list))],
        "roic": [to_pct(g(roic_list, i)) for i in range(len(roic_list))],
        "capex": [g(capex_list, i) for i in range(len(capex_list))],
        "sga": [g(sga_list, i) for i in range(len(sga_list))],
        "da": [g(da_list, i) for i in range(len(da_list))],
        "pe_ratio": [g(pe_list, i) for i in range(len(pe_list))],
        "pb_ratio": [g(pb_list, i) for i in range(len(pb_list))],
        "debt_ebitda": [g(debt_ebitda_list, i) for i in range(len(debt_ebitda_list))],
        "nd_ebitda": [g(nd_ebitda_list, i) for i in range(len(nd_ebitda_list))],
        "dividend_yield": [to_pct(g(dividend_yield_list, i)) for i in range(len(dividend_yield_list))],
        "payout_ratio": [to_pct(g(payout_ratio_list, i)) for i in range(len(payout_ratio_list))],
    }

    # 営業利益率の時系列が空で、計算可能な場合
    if not ts_data["op_margin"] and op_income and revenue:
        ts_data["op_margin"] = []
        for i in range(len(op_income)):
            oi = g(op_income, i)
            rev = g(revenue, i)
            if oi is not None and rev and rev != 0:
                ts_data["op_margin"].append(round(oi / rev * 100, 2))
            else:
                ts_data["op_margin"].append(None)

    # DuPont分解
    max_len = len(revenue) if revenue else 0
    net_margin_ts = []
    asset_turnover_ts = []
    fin_leverage_ts = []
    for i in range(max_len):
        ni = g(net_income, i)
        rev = g(revenue, i)
        ta = g(total_assets_list, i)
        eq = g(total_equity_list, i)
        if ni is not None and rev and rev != 0:
            net_margin_ts.append(round(ni / rev * 100, 2))
        else:
            net_margin_ts.append(None)
        if rev is not None and ta and ta != 0:
            asset_turnover_ts.append(round(rev / ta, 3))
        else:
            asset_turnover_ts.append(None)
        if ta is not None and eq and eq != 0:
            fin_leverage_ts.append(round(ta / eq, 3))
        else:
            fin_leverage_ts.append(None)

    ts_data["net_margin"] = net_margin_ts
    ts_data["asset_turnover"] = asset_turnover_ts
    ts_data["financial_leverage"] = fin_leverage_ts

    # 純利益率の分解時系列
    interest_burden_ts = []
    tax_burden_ts = []
    nonop_burden_ts = []
    for i in range(max_len):
        oi = g(op_income, i)
        pt = g(pretax_income_list, i)
        ni = g(net_income, i)
        ie = g(interest_exp_list, i)

        if oi is not None and oi != 0 and ie is not None:
            interest_burden_ts.append(round((oi + ie) / oi * 100, 2))
        else:
            interest_burden_ts.append(None)

        if oi is not None and ie is not None and (oi + ie) != 0 and pt is not None:
            nonop_burden_ts.append(round(pt / (oi + ie) * 100, 2))
        else:
            nonop_burden_ts.append(None)

        if pt is not None and pt != 0 and ni is not None:
            tax_burden_ts.append(round(ni / pt * 100, 2))
        else:
            tax_burden_ts.append(None)

    ts_data["interest_burden"] = interest_burden_ts
    ts_data["nonop_burden"] = nonop_burden_ts
    ts_data["tax_burden"] = tax_burden_ts

    # Metadata for frontend scaling
    ts_data["_source"] = "excel"
    ts_data["_unit_scale"] = "millions"  # Data is already in millions
    ts_data["_is_jpy"] = False  # Prevent double-scaling in smartFormat
    ts_data["_currency"] = currency

    return data, ts_data


# ---------- カスタム分析: 可視化可能データのスキャン ----------

# メトリクス定義: key, Excelラベル, シート名, カテゴリ, 単位, 表示名(ja), 表示名(en)
METRIC_CATALOG = [
    {"key": "revenue",          "sheet": "inc", "label": "Revenue",                    "cat": "growth",      "unit": "百万", "ja": "売上高",              "en": "Revenue"},
    {"key": "op_income",        "sheet": "inc", "label": "Operating Income",            "cat": "profitability","unit": "百万", "ja": "営業利益",            "en": "Operating Income"},
    {"key": "ebitda",           "sheet": "inc", "label": "EBITDA",                      "cat": "profitability","unit": "百万", "ja": "EBITDA",              "en": "EBITDA"},
    {"key": "net_income",       "sheet": "inc", "label": "Net Income",                 "cat": "profitability","unit": "百万", "ja": "純利益",              "en": "Net Income"},
    {"key": "eps",              "sheet": "inc", "label": "EPS (Basic)",                "cat": "growth",      "unit": "円",   "ja": "EPS",                 "en": "EPS"},
    {"key": "op_margin",        "sheet": "inc", "label": "Operating Margin",           "cat": "profitability","unit": "%",    "ja": "営業利益率",          "en": "Operating Margin"},
    {"key": "ebitda_margin",    "sheet": "inc", "label": "EBITDA Margin",              "cat": "profitability","unit": "%",    "ja": "EBITDAマージン",      "en": "EBITDA Margin"},
    {"key": "sga",              "sheet": "inc", "label": "Selling, General & Admin",   "cat": "other",       "unit": "百万", "ja": "販管費",              "en": "SG&A"},
    {"key": "da",               "sheet": "inc", "label": "Depreciation & Amortization","cat": "other",       "unit": "百万", "ja": "減価償却費",          "en": "D&A"},
    {"key": "cogs",             "sheet": "inc", "label": "Cost of Revenue",            "cat": "other",       "unit": "百万", "ja": "売上原価",            "en": "COGS"},
    {"key": "fcf",              "sheet": "cf",  "label": "Free Cash Flow",             "cat": "health",      "unit": "百万", "ja": "FCF",                 "en": "FCF"},
    {"key": "ocf",              "sheet": "cf",  "label": "Operating Cash Flow",        "cat": "health",      "unit": "百万", "ja": "営業CF",              "en": "Operating CF"},
    {"key": "capex",            "sheet": "cf",  "label": "Capital Expenditures",       "cat": "other",       "unit": "百万", "ja": "設備投資",            "en": "CapEx"},
    {"key": "investing_cf",     "sheet": "cf",  "label": "Investing Cash Flow",        "cat": "health",      "unit": "百万", "ja": "投資CF",              "en": "Investing CF"},
    {"key": "financing_cf",     "sheet": "cf",  "label": "Financing Cash Flow",        "cat": "health",      "unit": "百万", "ja": "財務CF",              "en": "Financing CF"},
    {"key": "total_assets",     "sheet": "bs",  "label": "Total Assets",               "cat": "health",      "unit": "百万", "ja": "総資産",              "en": "Total Assets"},
    {"key": "total_equity",     "sheet": "bs",  "label": "Shareholders Equity",        "cat": "health",      "unit": "百万", "ja": "自己資本",            "en": "Equity"},
    {"key": "total_debt",       "sheet": "bs",  "label": "Total Debt",                 "cat": "health",      "unit": "百万", "ja": "有利子負債",          "en": "Total Debt"},
    {"key": "cash",             "sheet": "bs",  "label": "Cash & Cash Equivalents",    "cat": "health",      "unit": "百万", "ja": "現金",                "en": "Cash"},
    {"key": "receivables",      "sheet": "bs",  "label": "Receivables",                "cat": "other",       "unit": "百万", "ja": "売上債権",            "en": "Receivables"},
    {"key": "inventory",        "sheet": "bs",  "label": "Inventory",                  "cat": "other",       "unit": "百万", "ja": "棚卸資産",            "en": "Inventory"},
    {"key": "current_assets",   "sheet": "bs",  "label": "Total Current Assets",       "cat": "health",      "unit": "百万", "ja": "流動資産",            "en": "Current Assets"},
    {"key": "current_liab",     "sheet": "bs",  "label": "Total Current Liabilities",  "cat": "health",      "unit": "百万", "ja": "流動負債",            "en": "Current Liabilities"},
    {"key": "pe_ratio",         "sheet": "rat", "label": "PE Ratio",                   "cat": "valuation",   "unit": "x",    "ja": "PER",                 "en": "P/E Ratio"},
    {"key": "pb_ratio",         "sheet": "rat", "label": "PB Ratio",                   "cat": "valuation",   "unit": "x",    "ja": "PBR",                 "en": "P/B Ratio"},
    {"key": "roe",              "sheet": "rat", "label": "Return on Equity (ROE)",     "cat": "performance", "unit": "%",     "ja": "ROE",                 "en": "ROE"},
    {"key": "roa",              "sheet": "rat", "label": "Return on Assets (ROA)",     "cat": "performance", "unit": "%",     "ja": "ROA",                 "en": "ROA"},
    {"key": "roic",             "sheet": "rat", "label": "Return on Invested Capital (ROIC)","cat": "performance","unit": "%","ja": "ROIC",                "en": "ROIC"},
    {"key": "dividend_yield",   "sheet": "rat", "label": "Dividend Yield",                "cat": "performance", "unit": "%",     "ja": "配当利回り",          "en": "Dividend Yield"},
    {"key": "payout_ratio",     "sheet": "rat", "label": "Payout Ratio",                  "cat": "performance", "unit": "%",     "ja": "配当性向",            "en": "Payout Ratio"},
    {"key": "current_ratio",    "sheet": "rat", "label": "Current Ratio",              "cat": "health",      "unit": "%",     "ja": "流動比率",            "en": "Current Ratio"},
    {"key": "quick_ratio",      "sheet": "rat", "label": "Quick Ratio",                "cat": "health",      "unit": "%",     "ja": "当座比率",            "en": "Quick Ratio"},
    {"key": "debt_ebitda",      "sheet": "rat", "label": "Debt/EBITDA",                "cat": "health",      "unit": "x",     "ja": "Debt/EBITDA",         "en": "Debt/EBITDA"},
    {"key": "nd_ebitda",        "sheet": "rat", "label": "Net Debt/EBITDA",            "cat": "valuation",   "unit": "x",     "ja": "Net Debt/EBITDA",     "en": "Net Debt/EBITDA"},
    {"key": "debt_fcf",         "sheet": "rat", "label": "Debt/FCF",                   "cat": "health",      "unit": "x",     "ja": "Debt/FCF",            "en": "Debt/FCF"},
    {"key": "ev",               "sheet": "rat", "label": "Enterprise Value",           "cat": "valuation",   "unit": "百万",  "ja": "EV",                  "en": "Enterprise Value"},
    {"key": "fixed_assets",     "sheet": "bs",  "label": "Property, Plant & Equipment","cat": "other",       "unit": "百万",  "ja": "有形固定資産",        "en": "PP&E"},
    {"key": "intangibles",      "sheet": "bs",  "label": "Goodwill and Intangibles",   "cat": "other",       "unit": "百万",  "ja": "のれん・無形資産",    "en": "Goodwill & Intangibles"},
    {"key": "net_debt",         "sheet": "bs",  "label": "Net Cash (Debt)",            "cat": "health",      "unit": "百万",  "ja": "ネットキャッシュ",    "en": "Net Cash (Debt)"},
    {"key": "long_term_assets", "sheet": "bs",  "label": "Total Long-Term Assets",     "cat": "other",       "unit": "百万",  "ja": "固定資産",            "en": "Long-Term Assets"},
]


def scan_available_metrics(filepath):
    """Excelファイルをスキャンし、可視化可能なメトリクス一覧を返す"""
    wb = _load_workbook(filepath)

    # 日本語縦型レイアウトの場合は専用処理
    if _is_japanese_vertical_layout(wb):
        return _scan_japanese_metrics(wb)

    sheet_map = {
        'inc': _find_sheet(wb, ['Income-Annual', 'Income Statement', 'Income', 'Export', 'income']),
        'bs':  _find_sheet(wb, ['Balance-Sheet-Annual', 'Balance Sheet', 'Balance', 'balance']),
        'cf':  _find_sheet(wb, ['Cash-Flow-Annual', 'Cash Flow', 'Cash Flow Statement', 'cashflow']),
        'rat': _find_sheet(wb, ['Ratios-Annual', 'Ratios', 'Financial Ratios', 'ratios']),
    }

    available = []
    for m in METRIC_CATALOG:
        ws = sheet_map.get(m['sheet'])
        if ws is None:
            continue
        # ファジーマッチで検索
        data, _, score = _fuzzy_get_row_data(ws, m['key'])
        if data:
            numeric = [v for v in data if isinstance(v, (int, float))]
            if numeric:
                available.append({
                    "key": m['key'],
                    "ja": m['ja'],
                    "en": m['en'],
                    "cat": m['cat'],
                    "unit": m['unit'],
                    "data_points": len(numeric),
                    "latest_value": numeric[0],
                })
    return available


def _scan_japanese_metrics(wb):
    """日本語縦型レイアウトからメトリクス一覧を返す"""
    data, _ = _parse_japanese_vertical(wb)
    catalog_map = {m['key']: m for m in METRIC_CATALOG}
    available = []

    # data辞書からリスト形式のデータを検出
    check_keys = {
        'revenue': data.get('revenue', []),
        'op_income': [data.get('op_income_val')],
        'net_income': [data.get('net_income_val')],
        'eps': data.get('eps', []),
        'ocf': data.get('operating_cf', []),
        'investing_cf': data.get('investing_cf', []),
        'financing_cf': data.get('financing_cf', []),
        'total_assets': [data.get('total_assets')],
        'total_equity': [data.get('total_equity')],
        'roe': data.get('roe', []),
        'roa': data.get('roa', []),
        'payout_ratio': [data.get('payout_ratio')],
    }

    for key, vals in check_keys.items():
        m = catalog_map.get(key)
        if m is None:
            continue
        numeric = [v for v in vals if v is not None and isinstance(v, (int, float))]
        if numeric:
            available.append({
                "key": m['key'],
                "ja": m['ja'],
                "en": m['en'],
                "cat": m['cat'],
                "unit": m['unit'],
                "data_points": len(numeric),
                "latest_value": numeric[0],
            })
    return available


def extract_custom_timeseries(filepath, selected_keys):
    """選択されたメトリクスの時系列データを返す"""
    wb = _load_workbook(filepath)

    # 日本語縦型レイアウトの場合
    if _is_japanese_vertical_layout(wb):
        _, ts_data = _parse_japanese_vertical(wb)
        result = {"dates": ts_data.get("dates", [])}
        for key in selected_keys:
            if key in ts_data:
                result[key] = ts_data[key]
        return result

    sheet_map = {
        'inc': _find_sheet(wb, ['Income-Annual', 'Income Statement', 'Income', 'Export', 'income']),
        'bs':  _find_sheet(wb, ['Balance-Sheet-Annual', 'Balance Sheet', 'Balance', 'balance']),
        'cf':  _find_sheet(wb, ['Cash-Flow-Annual', 'Cash Flow', 'Cash Flow Statement', 'cashflow']),
        'rat': _find_sheet(wb, ['Ratios-Annual', 'Ratios', 'Financial Ratios', 'ratios']),
    }

    # 日付はincシートから取得（ファジー）
    inc_ws = sheet_map.get('inc')
    dates, _, _ = _fuzzy_get_row_data(inc_ws, 'dates') if inc_ws else ([], None, 0.0)
    date_strs = [str(d)[:4] if d else "" for d in dates]

    result = {"dates": date_strs}
    catalog_map = {m['key']: m for m in METRIC_CATALOG}

    for key in selected_keys:
        m = catalog_map.get(key)
        if not m:
            continue
        ws = sheet_map.get(m['sheet'])
        if ws is None:
            continue
        data, _, _ = _fuzzy_get_row_data(ws, key)
        is_pct = m['unit'] == '%'
        if is_pct:
            data = [v * 100 if isinstance(v, (int, float)) else None for v in data]
        result[key] = data

    return result
